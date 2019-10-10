import io
import os
import openpyxl
import json

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pathlib import Path

from google.oauth2 import service_account
from google.cloud import vision
from google.cloud.vision import types
from google.protobuf.json_format import MessageToJson


def rect_in_rect(json_data, area):
	name_rect = area
	i = 0
	coords = tuple(json_data["vertices"])

	if (coords[0]['x'] >= name_rect[0][0]): 
		if (coords[0]['y'] >= name_rect[0][1]):
			if (coords[1]['x'] <= name_rect[1][0]):
				if (coords[1]['y'] >= name_rect[1][1]):
					if (coords[2]['x'] <= name_rect[2][0]):
						if (coords[2]['y'] <= name_rect[2][1]):
							if (coords[3]['x'] >= name_rect[3][0]):
								if (coords[3]['y'] <= name_rect[3][1]):
									return True

credentials = service_account.Credentials.from_service_account_file(
	'EspaceACE-OCR-399e82d0c778.json')

scoped_credentials = credentials.with_scopes(['https://www.googleapis.com/auth/cloud-platform'])

client = vision.ImageAnnotatorClient(credentials=credentials)

with io.open('sample.jpg', 'rb') as image_file:
	content = image_file.read()

image = vision.types.Image(content=content)

response = client.text_detection(image=image)

json_sample = MessageToJson(response, preserving_proto_field_name = True)

# with io.open('response.json', 'rb') as json_sample :
# 	content = json_sample.read()

# json_sample = json.loads(content)

customer_name = ""
caisse_amount = ""
mutuelle_amount = ""
ht_amount = ""
tva_amount = ""
ttc_amount = ""
avoir_or_fact = ""
invoice_num = ""

for json in json_sample["text_annotations"]:
	# Nom du client
	if (rect_in_rect(json["bounding_poly"], ((750,300),(1550,300),(1550,400),(750,400)))):
		customer_name += " {0}".format(json["description"])
	# Montant de la caisse
	if (rect_in_rect(json["bounding_poly"], ((1173,1811),(1272,1811),(1272,1851),(1173,1851)))):
		caisse_amount += "{0}".format(json["description"])
	# Montant de la mutuelle
	if (rect_in_rect(json["bounding_poly"], ((1467,1811),(1571,1811),(1571,1851),(1467,1851)))):
		mutuelle_amount += "{0}".format(json["description"])
	# Montant Hors Taxe
	if (rect_in_rect(json["bounding_poly"], ((1128,1886),(1221,1886),(1221,1936),(1128,1936)))):
		ht_amount += "{0}".format(json["description"])
	# Montant de TVA
	if (rect_in_rect(json["bounding_poly"], ((1269,1886),(1391,1886),(1391,1936),(1269,1936)))):
		tva_amount += "{0}".format(json["description"])
	# Montant TTC
	if (rect_in_rect(json["bounding_poly"], ((1430,1886),(1602,1886),(1602,1936),(1430,1936)))):
		ttc_amount += "{0}".format(json["description"])
	# Avoir ou facture ?
	if (rect_in_rect(json["bounding_poly"], ((14,597),(253,597),(253,661),(14,661)))):
		avoir_or_fact += "{0}".format(json["description"])
	# Num fact ?
	if (rect_in_rect(json["bounding_poly"], ((171,650),(355,650),(355,704),(171,704)))):
		invoice_num += "{0}".format(json["description"])	

print("Nom du client :", customer_name)
print("Total caisse :", caisse_amount)
print("Total mutuelle :", mutuelle_amount)
print("HT amount :", ht_amount)
print("TVA amount :", tva_amount)
print("TTC amount :", ttc_amount)
print("Avoir ou facture ?", avoir_or_fact)
print("Numero de facture :", invoice_num)

file_exist = Path("EguillesPerf-export.xlsx")

dest_filename = 'EguillesPerf-export.xlsx'

if file_exist.is_file():
	wb = load_workbook('EguillesPerf-export.xlsx')
	ws1 = wb.active
	lastrow = ws1.max_row
	if (lastrow > 0):
		lastrow = lastrow + 1

	ws1.cell(column=1, row=lastrow, value=avoir_or_fact)
	ws1.cell(column=2, row=lastrow, value=invoice_num)
	ws1.cell(column=3, row=lastrow, value=customer_name)
	ws1.cell(column=4, row=lastrow, value=caisse_amount)
	ws1.cell(column=5, row=lastrow, value=mutuelle_amount)
	ws1.cell(column=6, row=lastrow, value=ht_amount)
	ws1.cell(column=7, row=lastrow, value=tva_amount)
	ws1.cell(column=8, row=lastrow, value=ttc_amount)

	wb.save(filename = dest_filename)
else:
	wb = Workbook()
	ws1 = wb.active
	ws1.title = 'Ventes'
	lastrow = 1

	ws1.cell(column=1, row=lastrow, value="Type")
	ws1.cell(column=2, row=lastrow, value="Numero")
	ws1.cell(column=3, row=lastrow, value="Nom client")
	ws1.cell(column=4, row=lastrow, value="Part caisse")
	ws1.cell(column=5, row=lastrow, value="Part mutuelle")
	ws1.cell(column=6, row=lastrow, value="HT")
	ws1.cell(column=7, row=lastrow, value="TVA")
	ws1.cell(column=8, row=lastrow, value="TTC")

	ws1.cell(column=1, row=lastrow+1, value=avoir_or_fact)
	ws1.cell(column=2, row=lastrow+1, value=invoice_num)
	ws1.cell(column=3, row=lastrow+1, value=customer_name)
	ws1.cell(column=4, row=lastrow+1, value=caisse_amount)
	ws1.cell(column=5, row=lastrow+1, value=mutuelle_amount)
	ws1.cell(column=6, row=lastrow+1, value=ht_amount)
	ws1.cell(column=7, row=lastrow+1, value=tva_amount)
	ws1.cell(column=8, row=lastrow+1, value=ttc_amount)

	wb.save(filename = dest_filename)	
